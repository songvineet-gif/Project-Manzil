"""Turn Netlify form submissions into a formatted Excel workbook.

Usage: python3 to_excel.py subs.json "Manzil Properties - Enquiries.xlsx"
"""
import json, sys, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

src = sys.argv[1] if len(sys.argv) > 1 else 'subs.json'
out = sys.argv[2] if len(sys.argv) > 2 else 'enquiries.xlsx'
subs = json.load(open(src))

COLUMNS = [
    ('#',                       12),
    ('Received (UTC)',          19),
    ('Full name',               28),
    ('Phone',                   20),
    ('Email',                   30),
    ('Estimated property value',24),
    ('Notes',                   52),
]

wb = Workbook()
ws = wb.active
ws.title = 'Enquiries'

NAVY  = 'FF081226'
GOLD  = 'FFC99A3C'
LINE  = 'FFE6E8EC'
thin  = Side(style='thin', color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Title row
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
t = ws.cell(row=1, column=1, value='Manzil Properties — website enquiries')
t.font = Font(size=14, bold=True, color=NAVY)
t.alignment = Alignment(vertical='center')
ws.row_dimensions[1].height = 26

stamp = ws.cell(row=2, column=1,
                value='Exported %s UTC · %d enquir%s'
                      % (datetime.datetime.utcnow().strftime('%d %b %Y %H:%M'),
                         len(subs), 'y' if len(subs) == 1 else 'ies'))
stamp.font = Font(size=10, italic=True, color='FF5B6472')
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))

HEADER_ROW = 4
for i, (title, width) in enumerate(COLUMNS, start=1):
    c = ws.cell(row=HEADER_ROW, column=i, value=title)
    c.font = Font(bold=True, color='FFFFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=NAVY)
    c.alignment = Alignment(vertical='center', horizontal='left')
    c.border = border
    ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[HEADER_ROW].height = 22

def field(sub, key):
    for f in sub.get('ordered_human_fields', []):
        if f.get('name') == key:
            return f.get('value') or ''
    return (sub.get('data') or {}).get(key, '') or ''

# Newest first — the row you care about is the one that just came in
subs_sorted = sorted(subs, key=lambda s: s.get('created_at') or '', reverse=True)

for r, sub in enumerate(subs_sorted, start=HEADER_ROW + 1):
    ts = (sub.get('created_at') or '').replace('T', ' ').replace('Z', '')[:16]
    row = [r - HEADER_ROW, ts, field(sub, 'name'), field(sub, 'phone'),
           field(sub, 'email'), field(sub, 'value'), field(sub, 'message')]
    for i, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=val)
        c.alignment = Alignment(vertical='top', wrap_text=(i == 7))
        c.border = border
        if i == 1:
            c.font = Font(bold=True, color=NAVY)
        if i == 5 and val:
            c.hyperlink = 'mailto:%s' % val
            c.font = Font(color=GOLD, underline='single')
    ws.row_dimensions[r].height = 30

last = HEADER_ROW + max(len(subs_sorted), 1)
ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)
ws.auto_filter.ref = 'A%d:%s%d' % (HEADER_ROW, get_column_letter(len(COLUMNS)), last)

if not subs_sorted:
    e = ws.cell(row=HEADER_ROW + 1, column=1, value='No enquiries yet.')
    e.font = Font(italic=True, color='FF5B6472')

wb.save(out)
print('wrote %s — %d row(s)' % (out, len(subs_sorted)))
